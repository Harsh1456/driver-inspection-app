from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
import os
import uuid
import json
import io
from datetime import datetime
from werkzeug.utils import secure_filename
from PIL import Image  # Add this import

# Import modules
from config import Config
from database import db, UploadedFile, ReportPage, VehicleInspection, InspectionEdit  # Make sure all models are imported
from uploader import FileUploader
from classifier import RemarkClassifier
from extractor import TextExtractor
from dashboard import dashboard_bp

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(Config)

# Initialize components
db.init_app(app)
file_uploader = FileUploader(app.config['UPLOAD_FOLDER'], app.config['ALLOWED_EXTENSIONS'])

# Initialize AI components with debugging
classifier = None
text_extractor = None

print("=" * 50)
print("Initializing AI Components...")
print("=" * 50)

try:
    print(f"Attempting to load YOLO model from: {app.config['YOLO_MODEL_PATH']}")
    print(f"Model file exists: {os.path.exists(app.config['YOLO_MODEL_PATH'])}")
    
    # Add some debugging
    import torch
    print(f"PyTorch version: {torch.__version__}")
    print(f"CUDA available: {torch.cuda.is_available()}")
    
    classifier = RemarkClassifier(app.config['YOLO_MODEL_PATH'])
    print("✓ YOLO classifier initialized successfully")
except Exception as e:
    print(f"✗ Failed to initialize classifier: {str(e)}")
    import traceback
    traceback.print_exc()
    classifier = None

try:
    api_key = app.config.get('OPENAI_API_KEY')
    if api_key and not api_key.startswith('your-openai-api-key'):
        text_extractor = TextExtractor(api_key)
        print("✓ Text extractor initialized successfully")
    else:
        print("✗ OpenAI API key not configured properly")
        text_extractor = None
except Exception as e:
    print(f"✗ Failed to initialize text extractor: {str(e)}")
    text_extractor = None

print("=" * 50)

# Register blueprints
app.register_blueprint(dashboard_bp)

@app.route('/')
def home():
    """Redirect to dashboard"""
    return render_template('index.html')

@app.route('/upload')
def upload_page():
    """File upload interface"""
    return render_template('upload.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """API endpoint for file upload and processing"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Check if AI components are available
        if classifier is None:
            print("ERROR: Classifier is None at upload time")
            return jsonify({'success': False, 'error': 'YOLO classifier not available'}), 500
        if text_extractor is None:
            return jsonify({'success': False, 'error': 'Text extractor not available. Please check OpenAI API key configuration.'}), 500
        
        # Save uploaded file
        save_result = file_uploader.save_uploaded_file(file)
        if not save_result['success']:
            return jsonify({'success': False, 'error': save_result['error']}), 400
        
        # Process the file
        process_result = process_uploaded_file(save_result)
        
        return jsonify(process_result)
        
    except Exception as e:
        print(f"Upload error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def process_uploaded_file(file_info):
    """
    Process uploaded file through classification and extraction pipeline with batching
    """
    file_id = file_info['file_id']
    file_path = file_info['file_path']
    file_type = file_info['file_type']
    
    try:
        # Create file record in database
        file_record = UploadedFile(
            file_id=file_id,
            file_name=file_info['original_filename'],
            file_type=file_type,
            file_path=file_path
        )
        db.session.add(file_record)
        
        # Convert file to images
        images_dir = os.path.join(app.config['UPLOAD_FOLDER'], file_id)
        os.makedirs(images_dir, exist_ok=True)
        image_paths = []
        
        if file_type.lower() == 'pdf':
            image_paths = file_uploader.convert_pdf_to_images(file_path, images_dir)
        else:
            image_paths = file_uploader.process_single_image(file_path, images_dir)
        
        if not image_paths:
            raise Exception("Failed to convert file to images")
        
        # Extract header info from the first page
        if text_extractor:
            try:
                first_page_image = Image.open(image_paths[0])
                header_info = text_extractor.extract_header_info(first_page_image)
                
                if header_info['success']:
                    data = header_info['data']
                    inspection_record = VehicleInspection(
                        file_id=file_id,
                        carrier_name=data.get('carrier_name'),
                        location=data.get('location'),
                        inspection_date=data.get('date'),
                        inspection_time=data.get('time'),
                        truck_number=data.get('truck_number'),
                        odometer_reading=data.get('odometer')
                    )
                    db.session.add(inspection_record)
            except Exception as e:
                print(f"Header extraction failed: {str(e)}")

        # Process each page - FIRST PASS: Only classification
        pages_data = []
        pages_with_remarks = 0
        
        for i, image_path in enumerate(image_paths):
            page_id = str(uuid.uuid4())
            
            # Classify page
            classification_result = classifier.classify_image(image_path)
            
            has_remarks = classification_result['has_remarks']
            confidence = classification_result['confidence']
            bounding_boxes = classification_result['bounding_boxes']
            
            if has_remarks:
                pages_with_remarks += 1
            
            # Store page data for batch processing
            pages_data.append({
                'page_id': page_id,
                'page_number': i + 1,
                'image_path': image_path,
                'has_remarks': has_remarks,
                'confidence': confidence,
                'bounding_boxes': bounding_boxes,
                'extracted_text': "",
                'original_text': "",  # Store original extracted text
                'extraction_confidence': 0.0,
                'correction_applied': False  # Track if correction was applied
            })
        
        # SECOND PASS: Batch text extraction for pages with remarks
        if pages_with_remarks > 0:
            # print(f"Batch processing {pages_with_remarks} pages with remarks...")
            
            # Extract all remark regions first
            remark_images = []
            remark_page_indices = []
            
            for i, page_data in enumerate(pages_data):
                if page_data['has_remarks']:
                    remarks_image = classifier.extract_remarks_region(
                        page_data['image_path'], 
                        page_data['bounding_boxes']
                    )
                    if remarks_image:
                        remark_images.append(remarks_image)
                        remark_page_indices.append(i)
            
            # Batch extract text from all remark images
            if remark_images:
                batch_result = text_extractor.batch_extract_text_from_images(remark_images)
                
                # Update pages with extracted text
                for idx, page_index in enumerate(remark_page_indices):
                    if idx < len(batch_result['texts']):
                        extracted_text = batch_result['texts'][idx]
                        correction_applied = batch_result.get('correction_applied', [False] * len(batch_result['texts']))[idx]
                        improvement_score = batch_result.get('improvement_scores', [0.0] * len(batch_result['texts']))[idx]
                        
                        # If batch extraction returned "NO_HANDWRITING_DETECTED" but we know there are remarks,
                        # try individual extraction as fallback
                        if extracted_text == "NO_HANDWRITING_DETECTED":
                            print(f"Batch extraction failed for page {pages_data[page_index]['page_number']}, trying individual extraction...")
                            individual_result = text_extractor.extract_text_from_image(remark_images[idx])
                            if individual_result['success'] and individual_result['text'] != "NO_HANDWRITING_DETECTED":
                                pages_data[page_index]['extracted_text'] = individual_result.get('corrected_text', individual_result['text'])
                                pages_data[page_index]['original_text'] = individual_result['text']
                                pages_data[page_index]['extraction_confidence'] = individual_result['confidence']
                                pages_data[page_index]['correction_applied'] = individual_result.get('correction_applied', False)
                                pages_data[page_index]['improvement_score'] = individual_result.get('improvement_score', 0.0)
                            else:
                                pages_data[page_index]['extracted_text'] = "Unable to extract text"
                                pages_data[page_index]['extraction_confidence'] = 0.0
                        else:
                            pages_data[page_index]['extracted_text'] = extracted_text
                            pages_data[page_index]['original_text'] = batch_result.get('original_texts', [extracted_text] * len(batch_result['texts']))[idx]
                            pages_data[page_index]['extraction_confidence'] = batch_result['confidences'][idx]
                            pages_data[page_index]['correction_applied'] = correction_applied
                            pages_data[page_index]['improvement_score'] = improvement_score
        
        # Create database records
        pages_without_remarks = len(image_paths) - pages_with_remarks
        
        for page_data in pages_data:
            page_record = ReportPage(
                page_id=page_data['page_id'],
                file_id=file_id,
                page_number=page_data['page_number'],
                has_remarks=page_data['has_remarks'],
                extracted_text=page_data['extracted_text'],  # This will now store corrected text
                original_text=page_data.get('original_text', page_data['extracted_text']),  # Store original
                correction_applied=page_data.get('correction_applied', False),  # Track correction
                improvement_score=page_data.get('improvement_score', 0.0),  # Store improvement score
                confidence_score=page_data['extraction_confidence'] if page_data['has_remarks'] else page_data['confidence'],
                image_path=page_data['image_path'],
                bounding_boxes=json.dumps(page_data['bounding_boxes'])
            )
            db.session.add(page_record)
        
        # Calculate criticality
        total_pages = len(image_paths)
        if total_pages > 0:
            remarks_percentage = (pages_with_remarks / total_pages) * 100
            
            if remarks_percentage > 60:
                criticality = 'RED'
            elif remarks_percentage > 30:
                criticality = 'ORANGE'
            else:
                criticality = 'GREEN'
        else:
            criticality = 'GREEN'
        
        # Update file record with processing results
        file_record.total_pages = total_pages
        file_record.pages_with_remarks = pages_with_remarks
        file_record.pages_without_remarks = pages_without_remarks
        file_record.criticality_level = criticality
        
        # Commit all changes
        db.session.commit()
        
        return {
            'success': True,
            'file_id': file_id,
            'total_pages': total_pages,
            'pages_with_remarks': pages_with_remarks,
            'criticality': criticality,
            'message': 'File processed successfully'
        }
        
    except Exception as e:
        db.session.rollback()
        print(f"Error processing file: {str(e)}")
        traceback.print_exc()
        return {
            'success': False,
            'error': str(e)
        }

@app.route('/image/<file_id>/<int:page_number>')
def serve_image(file_id, page_number):
    """Serve images"""
    try:
        # Get page from database
        page = ReportPage.query.filter_by(file_id=file_id, page_number=page_number).first()
        if not page or not page.image_path:
            return "Image not found", 404
        
        # Check if image file exists
        if not os.path.exists(page.image_path):
            return "Image file not found", 404
            
        return send_file(page.image_path)
        
    except Exception:
        return "Image not found", 404

@app.route('/api/file/<file_id>/delete', methods=['DELETE'])
def delete_file(file_id):
    """Delete a file and all its associated images"""
    try:
        file = UploadedFile.query.get_or_404(file_id)
        
        # Delete associated images from filesystem
        for page in file.pages:
            if page.image_path and os.path.exists(page.image_path):
                try:
                    os.remove(page.image_path)
                except Exception:
                    pass
        
        # Delete the directory if empty
        file_dir = os.path.join(app.config['UPLOAD_FOLDER'], file_id)
        if os.path.exists(file_dir) and os.path.isdir(file_dir):
            try:
                if not os.listdir(file_dir):
                    os.rmdir(file_dir)
            except Exception:
                pass
        
        # Delete database records
        ReportPage.query.filter_by(file_id=file_id).delete()
        db.session.delete(file)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'File and associated images deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error deleting file: {str(e)}'
        }), 500

@app.route('/api/page/<page_id>/update-text', methods=['POST'])
def update_page_text(page_id):
    """Update extracted text for a page"""
    try:
        page = ReportPage.query.get_or_404(page_id)
        new_text = request.json.get('text', '')
        
        page.extracted_text = new_text
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Text updated successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error updating text: {str(e)}'
        }), 500

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """Export inspection data to Excel with signature"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.drawing.image import Image as ExcelImage
        import tempfile
        import base64
        import shutil
        
        file_ids = request.json.get('file_ids', [])
        
        if not file_ids:
            inspections = VehicleInspection.query.all()
        else:
            inspections = VehicleInspection.query.filter(VehicleInspection.file_id.in_(file_ids)).all()
        
        # Get the latest edit/signature for each file
        file_edits = {}
        for file_id in file_ids:
            edit = InspectionEdit.query.filter_by(file_id=file_id).order_by(InspectionEdit.edited_at.desc()).first()
            if edit:
                file_edits[file_id] = edit
        
        # Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inspection Reports"
        
        # Create header style
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = openpyxl.styles.PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write main headers
        headers = ['File ID', 'Carrier Name', 'Location', 'Inspection Date', 'Inspection Time', 
                   'Truck Number', 'Odometer Reading', 'Total Pages', 'Pages with Remarks', 
                   'Criticality Level', 'Processed Date']
        ws.append(headers)
        
        # Apply header styles
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Write data rows
        for insp in inspections:
            file_record = db.session.get(UploadedFile, insp.file_id)
            ws.append([
                insp.file_id,
                insp.carrier_name or 'N/A',
                insp.location or 'N/A',
                insp.inspection_date or 'N/A',
                insp.inspection_time or 'N/A',
                insp.truck_number or 'N/A',
                insp.odometer_reading or 'N/A',
                file_record.total_pages if file_record else 0,
                file_record.pages_with_remarks if file_record else 0,
                file_record.criticality_level if file_record else 'GREEN',
                insp.created_at.strftime('%Y-%m-%d %H:%M:%S') if insp.created_at else 'N/A'
            ])
        
        # Apply data row styles
        for row in ws.iter_rows(min_row=2, max_row=len(inspections) + 1, max_col=len(headers)):
            for cell in row:
                cell.border = border
                if cell.column == 10:  # Criticality column
                    if cell.value == 'RED':
                        cell.font = Font(color="DC2626", bold=True)
                    elif cell.value == 'ORANGE':
                        cell.font = Font(color="EA580C", bold=True)
                    else:
                        cell.font = Font(color="16A34A", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add signatures after 2 blank rows
        current_row = len(inspections) + 4  # 1 header + data rows + 2 blank rows
        
        # List to store temporary files for cleanup
        temp_files = []
        
        for file_id, edit in file_edits.items():
            # Add signature section header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            header_cell = ws.cell(row=current_row, column=1)
            header_cell.value = f"Authorization Signature - File: {file_id}"
            header_cell.font = Font(bold=True, size=14, color="1E3A8A")
            header_cell.alignment = Alignment(horizontal='left')
            current_row += 2
            
            # Add signer information
            ws.cell(row=current_row, column=1).value = "Inspector:"
            ws.cell(row=current_row, column=2).value = edit.signer_name
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            
            ws.cell(row=current_row + 1, column=1).value = "Title/Role:"
            ws.cell(row=current_row + 1, column=2).value = edit.signer_role or "Inspector"
            
            ws.cell(row=current_row + 2, column=1).value = "Signature Date:"
            ws.cell(row=current_row + 2, column=2).value = edit.signature_date or (edit.edited_at.strftime('%Y-%m-%d') if edit.edited_at else 'N/A')
            
            current_row += 4
            
            # Add signature image if available
            if edit.signature_data and edit.signature_data.strip():
                try:
                    # Extract base64 image data
                    if 'base64,' in edit.signature_data:
                        img_data = edit.signature_data.split('base64,')[1]
                        img_bytes = base64.b64decode(img_data)
                        
                        # Create a temporary file with a unique name
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.png', mode='wb') as tmp_file:
                            tmp_file.write(img_bytes)
                            tmp_file_path = tmp_file.name
                            temp_files.append(tmp_file_path)  # Store for later cleanup
                        
                        # Add image to Excel
                        img = ExcelImage(tmp_file_path)
                        img.width = 200
                        img.height = 80
                        ws.add_image(img, f'A{current_row}')
                        current_row += 8  # Move down for next section
                    else:
                        # Handle direct image data (if not base64 format)
                        ws.cell(row=current_row, column=1).value = f"Signature: {edit.signer_name}"
                        current_row += 2
                        
                except Exception as e:
                    print(f"Error adding signature image for file {file_id}: {str(e)}")
                    ws.cell(row=current_row, column=1).value = f"Signature: {edit.signer_name} (Image unavailable)"
                    current_row += 2
            
            # Add edited remarks
            if edit.edited_remarks and edit.edited_remarks.strip():
                current_row += 2
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
                ws.cell(row=current_row, column=1).value = "Edited Remarks:"
                ws.cell(row=current_row, column=1).font = Font(bold=True, color="374151")
                current_row += 1
                
                remarks_cell = ws.cell(row=current_row, column=1)
                remarks_cell.value = edit.edited_remarks
                remarks_cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[current_row].height = min(80, len(edit.edited_remarks) // 5)  # Dynamic height
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                
                current_row += 12  # Space for next file's signature
        
        # Add footer
        footer_row = current_row + 2
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
        footer_cell = ws.cell(row=footer_row, column=1)
        footer_cell.value = f"This report was generated by Inspection Processor System on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer_cell.font = Font(italic=True, color="6B7280")
        footer_cell.alignment = Alignment(horizontal='center')
        
        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Clean up temporary files AFTER saving
        for tmp_file_path in temp_files:
            try:
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)
            except Exception as e:
                print(f"Error cleaning up temp file {tmp_file_path}: {str(e)}")
        
        return jsonify({
            'success': True,
            'excel_data': base64.b64encode(output.getvalue()).decode(),
            'filename': f'inspection_reports_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        })
        
    except Exception as e:
        import traceback
        print(f"Excel export error: {str(e)}")
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/file/<file_id>/edit', methods=['POST'])
def save_edit(file_id):
    """Save report edits and signature"""
    try:
        data = request.json
        
        edit = InspectionEdit(
            file_id=file_id,
            signature_data=data.get('signature_data'),
            signature_type=data.get('signature_type'),
            signer_name=data.get('signer_name'),
            signer_role=data.get('signer_role'),
            signature_date=data.get('signature_date'),
            edited_remarks=data.get('edited_remarks'),
            original_remarks=data.get('original_remarks')
        )
        
        db.session.add(edit)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Report edits saved successfully',
            'edit_id': edit.id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error saving edits: {str(e)}'
        }), 500

@app.route('/api/file/<file_id>/edits', methods=['GET'])
def get_edits(file_id):
    """Get edit history for a file"""
    try:
        edits = InspectionEdit.query.filter_by(file_id=file_id).order_by(InspectionEdit.edited_at.desc()).all()
        return jsonify({
            'success': True,
            'edits': [edit.to_dict() for edit in edits]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error fetching edits: {str(e)}'
        }), 500

# Create tables when app starts
with app.app_context():
    try:
        db.create_all()
        print("✓ Database tables created successfully")
    except Exception as e:
        print(f"✗ Database creation failed: {e}")

if __name__ == '__main__':
    # Create upload directory if it doesn't exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    
    print("Starting server on http://localhost:5000")
    print(f"Static folder: {app.static_folder}")
    print(f"Template folder: {app.template_folder}")
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    
    # This is only for development
    app.run(debug=True, host='0.0.0.0', port=5000)
